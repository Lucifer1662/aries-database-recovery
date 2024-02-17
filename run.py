# Import necessary modules
from DB import DB
from CheckPoint import CheckPoint
from DirtyTable import DirtyTable
from TransactionTable import TransactionTable
import pandas as pd
import numpy as np
import tkinter as tk
from tkinter import messagebox
from tkinter import ttk
from math import isnan
import sys
import io
import openpyxl

def create_label_entry(parent, text, row, column):
    label = ttk.Label(parent, text=text, font=('Arial', 12))
    label.grid(row=row, column=column, sticky="e", padx=5, pady=5)
    entry = ttk.Entry(parent, font=('Arial', 12))
    entry.grid(row=row, column=column+1, padx=5, pady=5, sticky="ew")
    return entry

def create_example_labels(parent, example_inputs, row):
    for i, example_input in enumerate(example_inputs):
        label = ttk.Label(parent, text=example_input, font=('Arial', 10), foreground='gray')
        label.grid(row=row+i, column=2, sticky="w", padx=5, pady=2)

def display_excel_preview(file_path, parent, row):
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        max_row = sheet.max_row
        max_col = sheet.max_column

        columns = [str(sheet.cell(row=1, column=c).value) for c in range(1, max_col)]

        tree = ttk.Treeview(parent, columns=columns, show="headings", selectmode="browse")
        tree.grid(row=row, column=0, columnspan=2, padx=5, pady=5, sticky="nsew")

        for col in columns:
            tree.heading(col, text=col)

        for r in range(2, min(max_row, 5) + 1):
            row_data = [sheet.cell(row=r, column=c).value for c in range(1, max_col)]
            tree.insert("", "end", values=row_data)

    except Exception as e:
        messagebox.showerror("Error", f"Failed to load Excel file: {e}")

def redirect_stdout_to_tkinter(text_widget):
    import sys
    class StdoutRedirector(object):
        def __init__(self, text_widget):
            self.text_widget = text_widget
        def write(self, message):
            self.text_widget.insert("end", message)
            self.text_widget.see("end")

    sys.stdout = StdoutRedirector(text_widget)

class StdoutRedirector(io.StringIO):
    def __init__(self, text_widget):
        self.text_widget = text_widget
        io.StringIO.__init__(self)

    def write(self, s):
        self.text_widget.configure(state=tk.NORMAL)
        self.text_widget.insert(tk.END, s)
        self.text_widget.see(tk.END)  # Scroll to the end
        self.text_widget.configure(state=tk.DISABLED)

def redirect_stdout_to_tkinter(text_widget):
    sys.stdout = StdoutRedirector(text_widget)


number_of_lsn = 0
first_lsn = 0
flush_lsn = 0
    

# Define the main function
def main():
    # with open('file', 'w') as sys.stdout:
        number_of_lsn = lines_od_lsn_entry.get()
        first_lsn = first_lsn_entry.get()
        flush_lsn = flush_lsn_entry.get()

        try:
            number_of_lsn = int(number_of_lsn)
            first_lsn = int(first_lsn)
            flush_lsn = int(flush_lsn)
        except ValueError:
            messagebox.showerror("Error", "Please enter valid integer values for all inputs.")
            return
        
        # Load the Excel file into a DataFrame
        df = pd.read_excel('example.xlsx', usecols=['lsn', 'prev_lsn', 'transaction_id', 'type', 'page_id'])

        # Create an empty list to store dictionaries
        data_list = []

        # Iterate through each row in the DataFrame
        for index, row in df.iterrows():
            # Create a dictionary for the row
            row_dict = {
                'lsn': row['lsn'],
                'prev_lsn': row['prev_lsn'],
                'transaction_id': row['transaction_id'],
                'type': row['type'],
                'page_id': row['page_id']
            }
            
            # Append the dictionary to the list
            data_list.append(row_dict)

        # Initialize the database
        db = DB()
        
        # Iterate over the specified number of LSNs
        for i in range(number_of_lsn):
            # Retrieve values from the data_list dictionary
            lsn = data_list[i]["lsn"]
            prev_lsn = data_list[i]["prev_lsn"]
            transaction_id = data_list[i]["transaction_id"]
            type_ = data_list[i]["type"]
            page_id = data_list[i]["page_id"]

            # Check if any of the values are NaN and convert them accordingly
            lsn = int(lsn) if not isnan(lsn) else None
            prev_lsn = int(prev_lsn) if not isnan(prev_lsn) else None
            transaction_id = int(transaction_id) if not isnan(transaction_id) else None
            page_id = int(page_id) if not isnan(page_id) else None

            # Log the data into the database
            db.log.log(lsn, prev_lsn, transaction_id, type_, page_id, None, None)
        
        # Set the database's lastCheckPoint and perform crash recovery
        db.last_checkpoint = CheckPoint(begin=first_lsn, end=flush_lsn, dirty_page_table=DirtyTable(), transaction_table=TransactionTable())
        db.crash_recover()

        # Print the database
        print(db)
        # Disable the button after processing once
        process_button.config(state=tk.DISABLED)
        # Close the window

def style_widgets():
    style = ttk.Style()

    # Configure style for labels
    style.configure("TLabel", font=('Arial', 16), foreground="#333")

    # Configure style for buttons
    style.configure("TButton", font=('Arial', 12), foreground="white", background="#4CAF50")
    style.map("TButton",
              foreground=[('active', 'black')],
              background=[('active', '#45a049')])

    # Configure style for Treeview
    style.configure("Treeview.Heading", font=('Arial', 12), background="#f0f0f0")
    style.configure("Treeview", font=('Arial', 12), background="white", foreground="black", rowheight=25, fieldbackground="#d3d3d3")

root = tk.Tk()
root.title("Combined Application")

style_widgets()

# Configure root frame
root_frame = ttk.Frame(root)
root_frame.grid(row=0, column=0, padx=10, pady=10, sticky="nsew")

# Background Label for root_frame
background_label = tk.Label(root_frame, background='#a020f0')
background_label.place(relwidth=1, relheight=1)

# Frame for the input section
input_frame = ttk.Frame(root_frame)
input_frame.grid(row=0, column=0, padx=10, pady=10, sticky="nsew")

lines_od_lsn_entry = create_label_entry(input_frame, "Lines of lsn:", 0, 0)
first_lsn_entry = create_label_entry(input_frame, "First lsn:", 1, 0)
flush_lsn_entry = create_label_entry(input_frame, "Flush lsn:", 2, 0)

example_inputs = ["e.g., 7", "e.g., 1", "e.g., 7"]
create_example_labels(input_frame, example_inputs, row=0)

process_button = ttk.Button(input_frame, text="Process LSNs", command=main)
process_button.grid(column=1, row=3, columnspan=2, pady=20, sticky="ew")

# Frame for the Excel preview section
preview_frame = ttk.Frame(root_frame)
preview_frame.grid(row=0, column=1, padx=10, pady=10, sticky="nsew")

excel_file_path = "example.xlsx"
display_excel_preview(excel_file_path, preview_frame, row=0)

# Frame for the text widget section
text_frame = ttk.Frame(root_frame)
text_frame.grid(row=1, column=0, columnspan=2, padx=10, pady=10, sticky="nsew")

text_widget = tk.Text(text_frame, wrap="word", font=('Arial', 12))
text_widget.pack(fill="both", expand=True)

redirect_stdout_to_tkinter(text_widget)

# Configure column and row weights to make the window responsive
root_frame.grid_columnconfigure(0, weight=1)
root_frame.grid_columnconfigure(1, weight=1)
root_frame.grid_rowconfigure(0, weight=1)
root_frame.grid_rowconfigure(1, weight=1)

root.geometry("1420x780")
root.mainloop()